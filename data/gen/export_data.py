"""
PRECIPICE data export pipeline.
Reads PRECIPICE_COMPOUND_DATABASE.xlsx and PRECIPICE_ECONOMY_MODEL.xlsx,
runs structural assertions, and writes generated Luau modules to src/shared/Data/.

Usage: python data/gen/export_data.py
Run from the repo root.
"""

import sys
import io
import os
import openpyxl

# Force UTF-8 stdout on Windows so checkmark characters don't crash
if sys.platform == "win32" and hasattr(sys.stdout, "buffer"):
	sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from collections import defaultdict, deque

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CDB_PATH = os.path.join(REPO_ROOT, "data", "PRECIPICE_COMPOUND_DATABASE.xlsx")
ECO_PATH = os.path.join(REPO_ROOT, "data", "PRECIPICE_ECONOMY_MODEL.xlsx")
OUT_DIR = os.path.join(REPO_ROOT, "src", "shared", "Data")

GENERATED_HEADER = "-- GENERATED FILE — never hand-edit; run data/gen/export_data.py\n"

# ──────────────────────────────────────────────────────────────────────────────
# Load workbooks
# ──────────────────────────────────────────────────────────────────────────────

def load_workbooks():
	cdb = openpyxl.load_workbook(CDB_PATH, data_only=True)
	eco = openpyxl.load_workbook(ECO_PATH, data_only=True)
	return cdb, eco

# ──────────────────────────────────────────────────────────────────────────────
# Parse sheets
# ──────────────────────────────────────────────────────────────────────────────

def parse_compounds(cdb):
	"""Returns list of compound dicts."""
	ws = cdb["Compounds"]
	compounds = []
	for r in range(3, ws.max_row + 1):
		cid = ws.cell(r, 1).value
		if not cid:
			continue
		compounds.append({
			"id": cid,
			"name": ws.cell(r, 2).value,
			"tier": ws.cell(r, 3).value,
			"starter": ws.cell(r, 4).value == "YES",
			"baseValue": ws.cell(r, 6).value,
			"timeMins": ws.cell(r, 7).value,
			"passivePerMin": ws.cell(r, 8).value,
			"flavor": ws.cell(r, 12).value or "",
		})
	return compounds

def parse_tier_constants(cdb):
	ws = cdb["Tier_Constants"]
	tiers = []
	for r in range(3, ws.max_row + 1):
		t = ws.cell(r, 1).value
		if t is None:
			continue
		tiers.append({
			"tier": t,
			"timeMins": ws.cell(r, 2).value,
			"passivePerMin": ws.cell(r, 3).value,
			"valueMin": ws.cell(r, 4).value,
			"valueMax": ws.cell(r, 5).value,
		})
	return tiers

def parse_recipes(cdb):
	ws = cdb["Recipes"]
	recipes = []
	for r in range(3, ws.max_row + 1):
		rid = ws.cell(r, 1).value
		if not rid:
			continue
		recipes.append({
			"id": rid,
			"resultId": ws.cell(r, 2).value,
			"resultTier": ws.cell(r, 4).value,
			"input1": ws.cell(r, 5).value,
			"input2": ws.cell(r, 7).value,
			"pairType": ws.cell(r, 9).value,  # "simple" or "variable"
			"variablePairId": ws.cell(r, 10).value,
			"inputClass": ws.cell(r, 11).value,
		})
	return recipes

def parse_exotics(cdb):
	ws = cdb["Exotics"]
	exotics = []
	for r in range(4, ws.max_row + 1):
		eid = ws.cell(r, 1).value
		if not eid or not str(eid).startswith("EX_"):
			continue
		exotics.append({
			"id": eid,
			"inputA": ws.cell(r, 2).value,
			"inputB": ws.cell(r, 4).value,
			"name": ws.cell(r, 6).value,
			"initialValue": ws.cell(r, 7).value,
			"claimBonus": ws.cell(r, 8).value,
			"floorValue": ws.cell(r, 9).value,
		})
	return exotics

def parse_event_compounds(cdb):
	ws = cdb["Event_Compounds"]
	events = []
	for r in range(3, ws.max_row + 1):
		eid = ws.cell(r, 1).value
		if not eid or not str(eid).startswith("EV_"):
			continue
		events.append({
			"id": eid,
			"name": ws.cell(r, 2).value,
			"tier": ws.cell(r, 3).value,
			"timeMins": ws.cell(r, 4).value,
			"baseValue": ws.cell(r, 5).value,
			"blueprintFlux": ws.cell(r, 6).value,
			"input1": ws.cell(r, 7).value,
			"input2": ws.cell(r, 9).value,
		})
	return events

def parse_economy_params(eco):
	ws = eco["Params"]
	rank_scores = {}
	rank_thresholds = {}
	prestige_costs = {}
	reading = None
	for r in range(1, ws.max_row + 1):
		label = ws.cell(r, 1).value
		if label is None:
			continue
		if "Rank score per natural" in str(label):
			reading = "rank_scores"
			continue
		if "Rank thresholds" in str(label):
			reading = "rank_thresholds"
			continue
		if "Prestige" in str(label) and "cost" in str(label):
			reading = "prestige"
			continue
		if reading == "rank_scores" and str(label).startswith("T"):
			val = ws.cell(r, 2).value
			if val is not None and isinstance(val, (int, float)):
				rank_scores[str(label)] = int(val)
		elif reading == "rank_thresholds" and ws.cell(r, 2).value is not None:
			title = str(label)
			score = ws.cell(r, 2).value
			if isinstance(score, (int, float)):
				rank_thresholds[title] = int(score)
		elif reading == "prestige" and isinstance(label, (int, float)):
			cost = ws.cell(r, 2).value
			if isinstance(cost, (int, float)):
				prestige_costs[int(label)] = int(cost)
	return rank_scores, rank_thresholds, prestige_costs

def parse_contract_pool(eco):
	ws = eco["Contract_Pool"]
	contracts = []
	for r in range(4, ws.max_row + 1):
		cid = ws.cell(r, 1).value
		if not cid or not str(cid).startswith("C"):
			continue
		contracts.append({
			"id": cid,
			"name": ws.cell(r, 2).value,
			"difficulty": ws.cell(r, 3).value,
			"target": ws.cell(r, 4).value,
			"pelletReward": ws.cell(r, 5).value,
			"rankScore": ws.cell(r, 6).value,
			"filter": ws.cell(r, 7).value,
			"notes": ws.cell(r, 8).value,
		})
	return contracts

# ──────────────────────────────────────────────────────────────────────────────
# Structural assertions (abort on failure)
# ──────────────────────────────────────────────────────────────────────────────

TIER_COUNTS = {1: 30, 2: 30, 3: 30, 4: 25, 5: 20, 6: 15}
RECIPE_RESULT_COUNTS = {1: 25, 2: 40, 3: 50, 4: 45, 5: 30, 6: 15}
STARTER_IDS = {"T1_01", "T1_02", "T1_03", "T1_04", "T1_05"}

# Tier legality: which input-tier pairs may produce which result tier (doc §7, §11)
LEGAL_PAIRINGS = {
	# (min_tier, max_tier): allowed_result_tiers
	(1, 1): {1, 2},
	(1, 2): {3},
	(2, 2): {3},
	(2, 3): {4},
	(3, 3): {4},
	(3, 4): {5},
	(4, 5): {6},
	(5, 5): {6},
}

def assert_fail(msg):
	print(f"\nASSERTION FAILED: {msg}", file=sys.stderr)
	sys.exit(1)

def run_assertions(compounds, recipes, exotics, events, contracts):
	print("Running structural assertions...")
	cmap = {c["id"]: c for c in compounds}

	# 1) Compound counts
	total = len(compounds)
	if total != 150:
		assert_fail(f"Expected 150 compounds, got {total}")
	for tier, expected in TIER_COUNTS.items():
		actual = sum(1 for c in compounds if c["tier"] == tier)
		if actual != expected:
			assert_fail(f"Tier {tier}: expected {expected} compounds, got {actual}")
	print(f"  ✓ Compound counts: 150 total, {'/'.join(str(TIER_COUNTS[t]) for t in range(1,7))}")

	# 2) Recipe counts
	total_recipes = len(recipes)
	if total_recipes != 205:
		assert_fail(f"Expected 205 recipes, got {total_recipes}")
	for tier, expected in RECIPE_RESULT_COUNTS.items():
		actual = sum(1 for r in recipes if r["resultTier"] == tier)
		if actual != expected:
			assert_fail(f"Result tier {tier}: expected {expected} recipes, got {actual}")
	print(f"  ✓ Recipe counts: 205 total, {'/'.join(str(RECIPE_RESULT_COUNTS[t]) for t in range(1,7))}")

	# 3) Pair uniqueness and variable pair count
	pair_map = defaultdict(list)
	for recipe in recipes:
		i1, i2 = recipe["input1"], recipe["input2"]
		pair = tuple(sorted([i1, i2]))
		pair_map[pair].append(recipe)
	for pair, rlist in pair_map.items():
		if len(rlist) > 2:
			assert_fail(f"Pair {pair} has {len(rlist)} recipes (max 2)")
	doubles = sum(1 for rlist in pair_map.values() if len(rlist) == 2)
	if doubles != 17:
		assert_fail(f"Expected exactly 17 variable pairs (doubles), got {doubles}")
	print(f"  ✓ Pair uniqueness: all pairs ≤2 mappings, exactly 17 doubles")

	# 4) Variable pair IDs consistent
	variable_pair_ids = set()
	for recipe in recipes:
		if recipe["pairType"] != "simple":
			if not recipe["variablePairId"]:
				assert_fail(f"Recipe {recipe['id']} is non-simple but has no variablePairId")
			variable_pair_ids.add(recipe["variablePairId"])
	if len(variable_pair_ids) != 17:
		assert_fail(f"Expected 17 distinct variable pair IDs, got {len(variable_pair_ids)}")
	print(f"  ✓ Variable pair IDs: 17 distinct pairs")

	# 5) Tier legality for all 205 recipes
	for recipe in recipes:
		r_tier = recipe["resultTier"]
		i1 = cmap.get(recipe["input1"])
		i2 = cmap.get(recipe["input2"])
		if not i1:
			assert_fail(f"Recipe {recipe['id']}: unknown input1 {recipe['input1']}")
		if not i2:
			assert_fail(f"Recipe {recipe['id']}: unknown input2 {recipe['input2']}")
		t1, t2 = i1["tier"], i2["tier"]
		key = (min(t1, t2), max(t1, t2))
		allowed = LEGAL_PAIRINGS.get(key)
		if allowed is None or r_tier not in allowed:
			assert_fail(
				f"Recipe {recipe['id']}: {recipe['input1']}(T{t1}) + {recipe['input2']}(T{t2}) → "
				f"T{r_tier} is not a legal pairing"
			)
		# No same-compound authored recipe
		if recipe["input1"] == recipe["input2"]:
			assert_fail(f"Recipe {recipe['id']}: same-compound authored recipe ({recipe['input1']})")
	print("  ✓ Tier legality: all 205 mappings legal, zero same-compound authored recipes")

	# 6) T6 never an authored input; T1–T5 each used ≥1 time
	inputs_used = set()
	for recipe in recipes:
		inputs_used.add(recipe["input1"])
		inputs_used.add(recipe["input2"])
	for cid, c in cmap.items():
		if c["tier"] == 6 and cid in inputs_used:
			assert_fail(f"T6 compound {cid} used as authored recipe input")
	for tier in range(1, 6):
		tier_compounds = [c["id"] for c in compounds if c["tier"] == tier]
		used = [cid for cid in tier_compounds if cid in inputs_used]
		if not used:
			assert_fail(f"No T{tier} compound used as input")
	print("  ✓ Input coverage: T1–T5 all used; T6 never an authored input")

	# 7) BFS reachability from 5 starters
	reachable = set(STARTER_IDS)
	# Also reachable: all T1 compounds directly (extraction) — but recipes for non-starter T1
	# need T1+T1 inputs; we need at least the starters to discover them.
	# Build adjacency: pair → result
	recipe_graph = []
	for recipe in recipes:
		recipe_graph.append((recipe["input1"], recipe["input2"], recipe["resultId"]))

	changed = True
	while changed:
		changed = False
		for i1, i2, result in recipe_graph:
			if i1 in reachable and i2 in reachable and result not in reachable:
				reachable.add(result)
				changed = True

	unreachable = [c["id"] for c in compounds if c["id"] not in reachable]
	if unreachable:
		assert_fail(f"BFS: {len(unreachable)} compounds unreachable from starters: {unreachable[:5]}...")
	print(f"  ✓ BFS reachability: all 150 compounds reachable from 5 starters")

	# 8) Exotics: exactly 120, all T6+T6 pairs, unique names
	if len(exotics) != 120:
		assert_fail(f"Expected 120 Exotics, got {len(exotics)}")
	exotic_names = [e["name"] for e in exotics]
	if len(set(exotic_names)) != 120:
		assert_fail(f"Exotic names not all unique: {len(set(exotic_names))} distinct names")
	for e in exotics:
		for inp in [e["inputA"], e["inputB"]]:
			if inp not in cmap:
				assert_fail(f"Exotic {e['id']}: unknown input {inp}")
			if cmap[inp]["tier"] != 6:
				assert_fail(f"Exotic {e['id']}: input {inp} is not T6")
		if e["initialValue"] < 250000 or e["initialValue"] > 800000:
			assert_fail(f"Exotic {e['id']}: initial value {e['initialValue']} outside 250k–800k")
	print("  ✓ Exotics: 120 unique, all T6+T6 pairs, values 250k–800k")

	# 9) Contracts: 24 templates, rewards inside §19 bands
	if len(contracts) != 24:
		assert_fail(f"Expected 24 contract templates, got {len(contracts)}")
	REWARD_BANDS = {
		"Easy": (200, 500),
		"Medium": (500, 1500),
		"Hard": (1500, 3000),
	}
	for c in contracts:
		band = REWARD_BANDS.get(c["difficulty"])
		if not band:
			assert_fail(f"Contract {c['id']}: unknown difficulty {c['difficulty']}")
		lo, hi = band
		if not (lo <= c["pelletReward"] <= hi):
			assert_fail(
				f"Contract {c['id']}: reward {c['pelletReward']} outside {c['difficulty']} "
				f"band {lo}–{hi}"
			)
	easy = sum(1 for c in contracts if c["difficulty"] == "Easy")
	medium = sum(1 for c in contracts if c["difficulty"] == "Medium")
	hard = sum(1 for c in contracts if c["difficulty"] == "Hard")
	if (easy, medium, hard) != (8, 8, 8):
		assert_fail(f"Expected 8/8/8 Easy/Medium/Hard contracts, got {easy}/{medium}/{hard}")
	print(f"  ✓ Contracts: 24 templates (8/8/8), all rewards inside §19 bands")

	print("All assertions PASSED.")

# ──────────────────────────────────────────────────────────────────────────────
# Luau emitters
# ──────────────────────────────────────────────────────────────────────────────

def luau_string(s):
	if s is None:
		return '""'
	s = str(s).replace("\\", "\\\\").replace('"', '\\"')
	return f'"{s}"'

def emit_compounds(compounds):
	lines = [GENERATED_HEADER, "return {"]
	for c in compounds:
		lines.append(f'\t[{luau_string(c["id"])}] = {{')
		lines.append(f'\t\tid = {luau_string(c["id"])},')
		lines.append(f'\t\tname = {luau_string(c["name"])},')
		lines.append(f'\t\ttier = {c["tier"]},')
		lines.append(f'\t\tstarter = {str(c["starter"]).lower()},')
		lines.append(f'\t\tbaseValue = {c["baseValue"]},')
		lines.append(f'\t\ttimeMins = {c["timeMins"]},')
		lines.append(f'\t\tpassivePerMin = {c["passivePerMin"]},')
		lines.append(f'\t\tflavor = {luau_string(c["flavor"])},')
		lines.append("\t},")
	lines.append("}")
	return "\n".join(lines) + "\n"

def emit_recipes(recipes):
	lines = [GENERATED_HEADER, "return {"]
	for r in recipes:
		lines.append(f'\t[{luau_string(r["id"])}] = {{')
		lines.append(f'\t\tid = {luau_string(r["id"])},')
		lines.append(f'\t\tresultId = {luau_string(r["resultId"])},')
		lines.append(f'\t\tresultTier = {r["resultTier"]},')
		lines.append(f'\t\tinput1 = {luau_string(r["input1"])},')
		lines.append(f'\t\tinput2 = {luau_string(r["input2"])},')
		lines.append(f'\t\tpairType = {luau_string(r["pairType"])},')
		vp = r["variablePairId"]
		lines.append(f'\t\tvariablePairId = {luau_string(vp) if vp else "nil"},')
		lines.append(f'\t\tinputClass = {luau_string(r["inputClass"])},')
		lines.append("\t},")
	lines.append("}")
	return "\n".join(lines) + "\n"

def emit_exotics(exotics):
	lines = [GENERATED_HEADER, "return {"]
	for e in exotics:
		lines.append(f'\t[{luau_string(e["id"])}] = {{')
		lines.append(f'\t\tid = {luau_string(e["id"])},')
		lines.append(f'\t\tinputA = {luau_string(e["inputA"])},')
		lines.append(f'\t\tinputB = {luau_string(e["inputB"])},')
		lines.append(f'\t\tname = {luau_string(e["name"])},')
		lines.append(f'\t\tinitialValue = {e["initialValue"]},')
		lines.append(f'\t\tclaimBonus = {e["claimBonus"]},')
		lines.append(f'\t\tfloorValue = {e["floorValue"]},')
		lines.append("\t},")
	lines.append("}")
	return "\n".join(lines) + "\n"

def emit_event_compounds(events):
	lines = [GENERATED_HEADER, "return {"]
	for e in events:
		lines.append(f'\t[{luau_string(e["id"])}] = {{')
		lines.append(f'\t\tid = {luau_string(e["id"])},')
		lines.append(f'\t\tname = {luau_string(e["name"])},')
		lines.append(f'\t\ttier = {e["tier"]},')
		lines.append(f'\t\ttimeMins = {e["timeMins"]},')
		lines.append(f'\t\tbaseValue = {e["baseValue"]},')
		lines.append(f'\t\tblueprintFlux = {e["blueprintFlux"]},')
		lines.append(f'\t\tinput1 = {luau_string(e["input1"])},')
		lines.append(f'\t\tinput2 = {luau_string(e["input2"])},')
		lines.append("\t},")
	lines.append("}")
	return "\n".join(lines) + "\n"

def emit_tier_constants(tiers):
	lines = [GENERATED_HEADER, "return {"]
	for t in tiers:
		lines.append(f'\t[{t["tier"]}] = {{')
		lines.append(f'\t\ttier = {t["tier"]},')
		lines.append(f'\t\ttimeMins = {t["timeMins"]},')
		lines.append(f'\t\tpassivePerMin = {t["passivePerMin"]},')
		lines.append(f'\t\tvalueMin = {t["valueMin"]},')
		lines.append(f'\t\tvalueMax = {t["valueMax"]},')
		lines.append("\t},")
	lines.append("}")
	return "\n".join(lines) + "\n"

def emit_economy_params(rank_scores, rank_thresholds, prestige_costs):
	lines = [GENERATED_HEADER, "return {"]
	lines.append("\trankScorePerCompletion = {")
	for k, v in rank_scores.items():
		lines.append(f'\t\t[{luau_string(k)}] = {v},')
	lines.append("\t},")
	lines.append("\trankThresholds = {")
	for title, score in rank_thresholds.items():
		lines.append(f'\t\t[{luau_string(title)}] = {score},')
	lines.append("\t},")
	lines.append("\tprestigeCosts = {")
	for level, cost in sorted(prestige_costs.items()):
		lines.append(f'\t\t[{level}] = {cost},')
	lines.append("\t},")
	lines.append("}")
	return "\n".join(lines) + "\n"

def emit_contract_pool(contracts):
	lines = [GENERATED_HEADER, "return {"]
	for c in contracts:
		lines.append(f'\t[{luau_string(c["id"])}] = {{')
		lines.append(f'\t\tid = {luau_string(c["id"])},')
		lines.append(f'\t\tname = {luau_string(c["name"])},')
		lines.append(f'\t\tdifficulty = {luau_string(c["difficulty"])},')
		lines.append(f'\t\ttarget = {c["target"]},')
		lines.append(f'\t\tpelletReward = {c["pelletReward"]},')
		lines.append(f'\t\trankScore = {c["rankScore"]},')
		lines.append(f'\t\tfilter = {luau_string(c["filter"])},')
		lines.append(f'\t\tnotes = {luau_string(c["notes"])},')
		lines.append("\t},")
	lines.append("}")
	return "\n".join(lines) + "\n"

# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

def write_file(path, content):
	os.makedirs(os.path.dirname(path), exist_ok=True)
	with open(path, "w", encoding="utf-8") as f:
		f.write(content)
	print(f"  Wrote {os.path.relpath(path, REPO_ROOT)}")

def main():
	print(f"Loading workbooks from {os.path.relpath(CDB_PATH, REPO_ROOT)} and {os.path.relpath(ECO_PATH, REPO_ROOT)}...")
	cdb, eco = load_workbooks()

	compounds = parse_compounds(cdb)
	tier_constants = parse_tier_constants(cdb)
	recipes = parse_recipes(cdb)
	exotics = parse_exotics(cdb)
	events = parse_event_compounds(cdb)
	rank_scores, rank_thresholds, prestige_costs = parse_economy_params(eco)
	contracts = parse_contract_pool(eco)

	print(f"Parsed: {len(compounds)} compounds, {len(recipes)} recipes, "
		f"{len(exotics)} exotics, {len(events)} event compounds, {len(contracts)} contracts")

	run_assertions(compounds, recipes, exotics, events, contracts)

	print("\nWriting generated Luau modules...")
	write_file(os.path.join(OUT_DIR, "Compounds.luau"), emit_compounds(compounds))
	write_file(os.path.join(OUT_DIR, "Recipes.luau"), emit_recipes(recipes))
	write_file(os.path.join(OUT_DIR, "Exotics.luau"), emit_exotics(exotics))
	write_file(os.path.join(OUT_DIR, "EventCompounds.luau"), emit_event_compounds(events))
	write_file(os.path.join(OUT_DIR, "TierConstants.luau"), emit_tier_constants(tier_constants))
	write_file(os.path.join(OUT_DIR, "EconomyParams.luau"), emit_economy_params(rank_scores, rank_thresholds, prestige_costs))
	write_file(os.path.join(OUT_DIR, "ContractPool.luau"), emit_contract_pool(contracts))

	print("\nDone. All files generated successfully.")

if __name__ == "__main__":
	main()
